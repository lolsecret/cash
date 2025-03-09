from typing import Optional, Set
import os.path
import re
import logging
import time
from collections import OrderedDict
from datetime import datetime
from zipfile import ZipFile, ZipInfo

import urllib3
import requests
import xmltodict
from django.conf import settings
from openpyxl import load_workbook
from django.db import transaction

from . import BlackListReason, BlackListSource
from .models import BlackListMember

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

logger = logging.getLogger(__name__)


def get_filename_zip(value):
    if isinstance(value, str):
        if matched := re.search(r'filename=[\'"]?([^\'"]+)', value):
            return matched.group(1)
    return 'request.zip'


def ips_download_file(url):
    # NOTE the stream=True parameter below
    with requests.get(url, stream=True, verify=False) as response:
        response.raise_for_status()

        filename = get_filename_zip(response.headers.get('content-disposition'))
        local_filename = os.path.join(settings.MEDIA_ROOT, filename)

        with open(local_filename, 'wb') as f:
            for chunk in response.iter_content(chunk_size=8192):
                f.write(chunk)

        return local_filename


def extract_excel_ips(filename_zip: str):
    with ZipFile(filename_zip, mode="r") as archive:
        if not len(archive.filelist):
            logger.error("extract_excel: %s is empty", filename_zip)
            raise Exception("%s is empty" % filename_zip)

        file: ZipInfo = archive.filelist[0]
        logger.info("archive.extract: %s", archive.extract(file.filename, path=settings.MEDIA_ROOT))
        return os.path.join(settings.MEDIA_ROOT, file.filename)


@transaction.atomic
def import_blacklist_from_xml(xml: str):
    records: OrderedDict = xmltodict.parse(xml)["xml"]["persons"]["person"]

    BlackListMember.objects.filter(source=BlackListSource.EGOV).delete()

    to_create = []
    iins: Set[str] = set()

    for record in records:
        iin = record["iin"]
        if iin in iins:
            continue
        else:
            iins.add(iin)

        if record["birthdate"]:
            birthday: Optional[datetime] = datetime.strptime(
                record["birthdate"], "%d.%m.%Y"
            )
        else:
            birthday = None

        to_create.append(
            BlackListMember(
                iin=iin,
                first_name=record["fname"],
                last_name=record["lname"],
                middle_name=record["mname"],
                birthday=birthday,
                note=record["note"],
                reason=BlackListReason.AML,
                source=BlackListSource.EGOV,
            )
        )

    created = BlackListMember.objects.bulk_create(to_create)
    return f"Loaded {len(created)} members from EGOV"


class ReadExcelIPs:
    BIN = 'БИН'
    # noinspection SpellCheckingInspection
    HEADER_NAMES = (
        (BIN, 'iin'),
        ('Полное наименование', 'name'),
        ('Дата регистрации', 'date_reg'),
        ('КАТО', 'kato_code'),
        ('Басшының ТАӘ, ФИО руководителя', 'full_name'),
    )

    def __init__(self, file) -> None:
        self.file = file
        self.results = []
        self.header_founded = {}

        self.workbook = None

    def open(self):
        self.workbook = load_workbook(filename=self.file, read_only=True)
        logger.info("ReadExcelIPs.open: workbook loaded")

    def close(self):
        self.workbook.close()
        logger.info("ReadExcelIPs.close: workbook closed")

    def __enter__(self):
        self.open()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()

    def find_by_name(self, value: str):
        return next((field for (name, field) in self.HEADER_NAMES if name == value.strip()), None)

    def find_header(self, rows):  # noqa
        for row in rows:
            if row[0].value and row[0].value.strip() == self.BIN:
                for index, cell in enumerate(row):  # noqa
                    row_founded = self.find_by_name(cell.value)
                    if row_founded:
                        self.header_founded[row_founded] = index
                return True
        return False

    def parse(self):
        start_time = time.perf_counter()

        for sheet_name in self.workbook.sheetnames:
            logger.info("ReadExcelIPs.parse: open sheet_name: %s", sheet_name)
            worksheet = self.workbook[sheet_name]
            worksheet.reset_dimensions()

            rows = worksheet.rows

            if self.find_header(rows):
                for row in rows:
                    if row[0].value:
                        item = {}
                        for key, idx in self.header_founded.items():
                            item[key] = row[idx].value

                        self.results.append(item)

        logger.info("ReadExcelIPs.parse: runtime %s", round(time.perf_counter() - start_time, 3))
